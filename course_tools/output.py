from __future__ import annotations

import os
import sys
from typing import TYPE_CHECKING, TypeVar, cast

import numpy as np


if TYPE_CHECKING:
    from collections.abc import Mapping

    from openpyxl import Worksheet

    from course_tools.data import Database, Student


from .grade_tools import format_frac


T = TypeVar("T")


try:
    from vt100 import c
except ImportError:
    def c(name):
        return ""

    from warnings import warn
    warn("vt100 module not found--colored output disabled", stacklevel=2)


def _assert_not_none(val: T | None) -> T:
    assert val is not None
    return val


def print_scales(database: Database):
    rules = database.course_rules
    for name, values in sorted(rules["SCALE_CUTOFFS"].items()):
        print("-"*75)
        print("SCALE: %s" % name)
        print("-"*75)
        for grade, points in zip(rules["LETTER_GRADES"], values, strict=False):
            print("%s\t >= %s" % (grade, points))


def print_warnings(database: Database, warn_level: int):
    for student in database.students.values():
        for severity, ln in student.log:
            if severity >= warn_level:
                print("*** %s: %s" % (student.network_id, ln), file=sys.stderr)


def find_student(database: Database, search_term: str):
    try:
        return database.students[search_term]
    except KeyError:
        pass

    values: list[Student] = []
    for student in database.students.values():
        if search_term.lower() in (
                student.network_id
                + "|" + getattr(student, "last_name", "")
                + "|" + getattr(student, "first_name", "")).lower():
            values.append(student)

    if not values:
        raise ValueError("no student found for '%s'" % search_term)
    if len(values) > 1:
        for student in values:
            print("%s: %s, %s" % (
                student.network_id, student.last_name, student.first_name))
        raise ValueError(
            "more than one student found for '%s'" % search_term)
    return values[0]


def print_student_report(database: Database, search_term: str):
    student = find_student(database, search_term)

    print("-"*75)
    print(
        "%s, %s (%s)"
        % (student.last_name, student.first_name, student.network_id))
    print("-"*75)
    print("UIN: %s" % student.university_id)
    print("Section: %s" % student.section)
    print(f"Credit hours: {student.credit_hours}")
    print("Standing: %s" % student.standing)
    print("Scale: %s" % database.course_rules["GET_SCALE"](student))
    for severity, ln in student.log:
        if severity >= 4:
            ln = c("bright red") + "/!\\ " + ln + c("normal")

        print(ln)

    print("-"*75)
    print("Grade: %s" % format_frac(student.grade))
    print("Rounded grade: %s" % student.rounded_grade)
    print("Letter grade: %s" % student.letter_grade)
    print("-"*75)


def print_grade_list(database: Database):
    sections = {
        str(getattr(student, "section", None))
        for student in database.students.values()}

    for section in sorted(sections):
        print("-"*75)
        print("SECTION %s" % section)
        print("-"*75)

        students = sorted(
            (student
                for student in database.students.values()
                if str(getattr(student, "section", None)) == section
                if student.grade is not None),
            # key=lambda student: student.last_name
            key=lambda student: student.rounded_grade,
            )

        for student in students:
            print(
                "%-10s %-6s %-20s %-20s %-10s %-3s %.1f"
                % (student.network_id,
                    str(getattr(student, "standing", None))[:6],
                    student.last_name,
                    student.first_name,
                    str(getattr(student, "university_id", None)),
                    student.letter_grade,
                    student.rounded_grade))


def plot_histogram(database: Database, differentiated: bool):
    sections = sorted(
            {str(getattr(student, "section", None))
            for student in database.students.values()})
    standings = sorted(
            {str(getattr(student, "standing", None))
            for student in database.students.values()})

    dataset_names = []
    datasets = []

    if differentiated:
        for standing in standings:
            for section in sections:
                students = [
                    student
                    for student in database.students.values()
                    if str(getattr(student, "section", None)) == section
                    if str(getattr(student, "standing", None)) == standing]

                if not students:
                    continue

                dataset = [
                    100*student.grade for student in students
                    if student.grade is not None]
                datasets.append(dataset)
                dataset_names.append("%s - %s" % (standing, section))
    else:
        datasets = [[100*student.grade for student in database.students.values()]]
        dataset_names = ["Everybody"]

    import matplotlib.pyplot as pt
    for name, ds in zip(dataset_names, datasets, strict=False):
        print("%s: mean: %.2f - stddev: %.2f (n=%d)" % (
            name, np.average(ds), np.std(ds), len(ds)))

    pt.hist(datasets, label=dataset_names, bins=15, histtype="barstacked")
    pt.legend(loc="best")
    pt.show()


def print_letter_histogram(database):
    letter_count = {}

    for student in database.students.values():
        ltr = student.letter_grade
        letter_count[ltr] = letter_count.get(ltr, 0) + 1

    for ltr in database.course_rules["LETTER_GRADES"]:
        count = letter_count.get(ltr, 0)
        print("%-3s : % 4d : %s" % (ltr, count, count*"#"))


def print_emails(database: Database, email_suffix="") -> None:
    students = sorted(
        database.students.values(),
        key=lambda student: cast("str", student.network_id))

    for student in students:
        print(f"{student.network_id}{email_suffix}")


def print_roster_csv(database: Database) -> None:
    students = sorted(
        database.students.values(),
        key=lambda student: cast("str", student.network_id))

    import csv
    writer = csv.writer(sys.stdout)
    writer.writerow(["NetID", "First Name", "Last Name"])
    for student in students:
        writer.writerow([student.network_id, student.first_name, student.last_name])


def print_banner_csv(database: Database) -> None:
    students = sorted(
        database.students.values(),
        key=lambda student: _assert_not_none(student.network_id))

    for student in students:
        if not student.university_id:
            print(
                "*** %s (Section %s) does not have a university_id"
                % (student.network_id, getattr(student, "section", None)),
                file=sys.stderr)
        else:
            print("%s,%-3s    # %s" % (
                student.university_id, student.letter_grade,
                student.network_id))


def _get_ws_column_headers(ws: Worksheet) -> Mapping[str, int]:
    result = {}
    i = 1
    while h := ws.cell(row=1, column=i).value:
        result[h] = i
        i += 1

    return result


def update_banner_xlsx(database: Database, filename: str) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    ws = wb.active

    headers = _get_ws_column_headers(ws)
    student_id_col = headers["Student ID"]
    grade_col = headers["Final Grade"]

    sid_to_row = {}
    i = 2
    while sid := ws.cell(row=i, column=student_id_col).value:
        sid_to_row[sid] = i
        i += 1

    for student in database.students.values():
        if not student.university_id:
            print(
                f"*** {student.network_id} "
                f"(Section {getattr(student, 'section', None)}) "
                "does not have a university_id",
                file=sys.stderr)
        elif student.university_id not in sid_to_row:
            print(
                f"*** {student.network_id} "
                f"(Section {getattr(student, 'section', None)}) "
                "is not in template XLS",
                file=sys.stderr)
        else:
            ws.cell(
                    row=sid_to_row[student.university_id],
                    column=grade_col,
                    value=student.letter_grade)

    nm, ext = os.path.splitext(filename)
    wb.save(f"{nm}-updated{ext}")


def print_relate_csv(database: Database) -> None:
    students = sorted(
        database.students.values(),
        key=lambda student: _assert_not_none(student.network_id))

    for student in students:
        if student.rounded_grade is not None:
            print("%s,%.2f,Letter grade: %s" % (
                student.network_id,
                student.rounded_grade,
                student.letter_grade))


def print_preliminary_relate_csv(database: Database):
    students = sorted(
        database.students.values(),
        key=lambda student: student.network_id)

    import csv
    writer = csv.writer(sys.stdout, quoting=csv.QUOTE_MINIMAL)
    for student in students:
        writer.writerow([
            student.network_id,
            "%.2f" % (100*student.grade),
            ""])


def print_relate_not_in_roster_query(database: Database, email_suffix: str) -> None:
    students = sorted(
        database.students.values(),
        key=lambda student: cast("str", student.network_id))

    pos_query = " or ".join(
            f"email:{student.network_id}{email_suffix}"
            for student in students)

    print(f"role:student and status:active and not ({pos_query})")
